from openpyxl import load_workbook
from openpyxl import Workbook

# The format of a MyPlate website's detailed export is:
# Date: | <date>
# 
# Meals
# <Meal header>
# <Meal rows>
#
#
# Fitness
# <Fitness header>
# <Meal rows>
#
#
# Totals:
# <indent with 4 empty cells> | <Totals header>
# <indent with 4 empty cells> | <Totals row>
# <indent with 4 empty cells> | <Totals calories summary header> | <Totals calories summary value>
#
# <repeat with next "Date:">
# Weight
#
# Date: | <date>
# Weight | <weight>
# <repeat with next "Date:" and "Weight">
#
# Water
# <Water header>
# <date> | <Water intake amount>


def init_sheets(t_meals_sheet, t_fitness_sheet, t_totals_sheet):
    """
    Insert headers into the sheets that will be used for target of split

    :param t_meals_sheet: split target sheet for Meals section
    :param t_fitness_sheet: split target sheet for Fitness section
    :param t_totals_sheet: split target sheet for Totals section
    """
    header = ["Date", "Meal", "Item Brand", "Item Name", "Your Servings", "Your Total Calories", "Your Total Sugars",
              "Your Total Carbs", "Your Total Fats", "Your Total Protein", "Your Total Cholesterol",
              "Your Total Sodium", "Your Total Dietary Fiber", "Calories", "Sugars", "Carbs", "Fats", "Protein",
              "Cholesterol", "Sodium", "Dietary Fiber"]
    t_meals_sheet.append(header)

    header = ["Date", "Exercise Done", "Minutes", "Calories Burned", "Heart Rate", "Distance"]
    t_fitness_sheet.append(header)

    header = ["Date", "Calories", "Sugars", "Carbohydrates", "Fat", "Protein", "Cholesterol", "Sodium",
              "Dietary Fiber", "Calories Allowed", "Calories Consumed", "Calories Burned", "Net Calories", "Weight",
              "Water"]
    t_totals_sheet.append(header)

    return t_meals_sheet, t_fitness_sheet, t_totals_sheet


def extract_meals_fitness(t_main_sheet, t_target_sheet, t_row, t_date):
    """
    Starting from t_row, extract a certain number of rows (until the next blank line) from t_main_sheet, and insert
    them into t_target_sheet. Also prepend the date for that section to each row.

    :param t_main_sheet: source sheet for where to extract appropriate section from
    :param t_target_sheet: split target sheet for either Meals or Fitness section
    :param t_row: row number to start extraction from (is a header row so actually row+1)
    :param t_date: date for this section which is being extracted
    """

    pass


def extract_totals(t_main_sheet, t_totals_sheet, t_row, t_cur_date):
    """

    :param t_main_sheet:
    :param t_totals_sheet:
    :param t_row:
    :param t_cur_date:
    """
    pass


def insert_weight(t_main_sheet, t_totals_sheet, t_row, t_cur_date):
    """
    Insert weights into the totals table. There are gaps (depending on when user records their weight), leave nulls in
    place.

    :param t_main_sheet:
    :param t_totals_sheet:
    :param t_row:
    :param t_cur_date:
    """
    pass


def insert_water(t_main_sheet, t_totals_sheet, t_row):
    """
    Insert water intake into the totals table. There are gaps (depending on when user records their water intake),
    leave nulls in place.

    :param t_main_sheet:
    :param t_totals_sheet:
    :param t_row:
    """
    # Note the date format is YYYY-MM-DD but the date in totals_sheet is written out, with the th and rd endings
    #  like October 14th, 2019
    pass


if __name__ == "__main__":
    # Load existing workbook from MyPlate website's export
    # TODO: Take file input from command line instead of hardcoding this
    main_workbook = load_workbook(filename="MyPlate-Export-2019-10-14_detailed.xlsx")
    # Create new workbooks for data that will be split
    meals = Workbook()
    fitness = Workbook()
    totals = Workbook()
    # Filenames to use for these workbooks when saving
    meals_fname = "meals.xlsx"
    fitness_fname = "fitness.xlsx"
    totals_fname = "totals.xlsx"

    # Set main sheet as active
    main_sheet = main_workbook.active
    meals_sheet = meals.active
    fitness_sheet = fitness.active
    totals_sheet = totals.active

    # Maximum number of rows in main worksheet
    #max_rows = len(main_sheet.rows)
    # Tracker for current row, incremented as needed in while loop. Start at 1 is there are any rows to process.
    if main_sheet.max_row != 0:
        row = 1
    else:
        row = 0
    # Tracker for current date
    cur_date = None

    # TODO: instead of "state" tracking, calculate the number of rows from the current row until the blank line and
    #  extract those (adding a date field in beginning, including to header), skipping header (write that separately?)
    # TODO: maybe skip the above number of iterations with next (see
    #  https://stackoverflow.com/questions/22295901/skip-multiple-iterations-in-loop ) but may get messy. Or do you
    #  need a for loop at all? Can also do while loop on a counter and add large section number to counter to kind of
    #  skip iterations. Would be cleaner/faster.
    # TODO: reformat the totals section as well so it's all one line with a date
    # for idx, row in enumerate(t_main_sheet.iter_rows(values_only=True)):

    # Initialize each target/split sheet with appropriate final headers
    init_sheets(meals_sheet, fitness_sheet, totals_sheet)

    # Loop until current row is the sheet's last row
    while row != main_sheet.max_row:
        # Check if current row is for Date section
        if main_sheet[row][0] == "Date:":
            cur_date = main_sheet[row][1]
            row += 1
            # Handle case at end of sheet where weights are listed depending on date, insert weight as column to totals
            if main_sheet[row][0] == "Weight":
                insert_weight(main_sheet, totals_sheet, row, cur_date)
                row += 1
        # Check if current row is for Meals section
        elif main_sheet[row][0] == "Meals":
            end_row = extract_meals_fitness(main_sheet, meals_sheet, row, cur_date)
            row += (end_row - row)
        # Check if current row is for Fitness section
        elif main_sheet[row][0] == "Fitness":
            end_row = extract_meals_fitness(main_sheet, fitness_sheet, row, cur_date)
            row += (end_row - row)
        # Check if current row is for Totals section
        elif main_sheet[row][0] == "Totals:":
            end_row = extract_totals(main_sheet, totals_sheet, row, cur_date)
            row += (end_row - row)
        # Check if current row is for Water section at end of sheet (iterate through date rows)
        elif main_sheet[row][0] == "Water":
            end_row = insert_water(main_sheet, totals_sheet, row)
            row += (end_row - row)
        # Otherwise ignore line
        else:
            row += 1

    # Write out final split sheets
    meals.save(meals_fname)
    fitness.save(fitness_fname)
    totals.save(totals_fname)
